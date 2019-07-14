#!/usr/bin/python3
# -*- coding: utf-8 -*-

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font,  Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


class Coordinates:
    def __init__(self, min_x, max_x, min_y, max_y, name, description, map_file):
        self.min_x = min_x
        self.max_x = max_x
        self.min_y = min_y
        self.max_y = max_y
        self.name = name
        self.description = description
        self.map_file = map_file

    def query_part(self):
        if self.min_x:
            return 'x >= {min_x} AND x <= {max_x} AND y >= {min_y} AND y <= {max_y}'\
                    .format(min_x=self.min_x, max_x=self.max_x, min_y=self.min_y,
                            max_y=self.max_y)
        else:
            return ''


class DateRange:
    def __init__(self, start, end, description):
        self.start = start
        self.end = end
        self.description = description

    def query_part(self):
        if self.start:
            return ' AND prediction_date >= {x} AND prediction_date <= {y}'\
                .format(x=self.start, y=self.end)
        else:
            return ''

    def map_dir_part(self):
        if self.start:
            return self.description.replace(' do ', '_to_')
        else:
            return 'all_data'


class Prediction:
    def __init__(self, start, end):
        self.start = start
        self.end = end

    def query_part(self):
        if self.start:
            return ' AND prediction_length >= {start} AND prediction_length <= {end}'\
                .format(start=self.start, end=self.end)
        else:
            return ''

    def description(self):
        if self.start:
            return '{start}h-{end}h'.format(start=self.start, end=self.end)
        else:
            return 'Wszystko'


def execute_query(cur, reg, dt_range, query_field, pred):
    query = ('SELECT COUNT(*), AVG({field}), AVG(ABS({field})), STDDEV_POP({field}) FROM fact_prediction' +
             ' WHERE {region_query}{date_range_query}{prediction_query}')\
        .format(field=query_field,
                region_query=reg.query_part(),
                date_range_query=dt_range.query_part(),
                prediction_query=pred.query_part())
    if query.endswith('WHERE '):
        query = query[:-6]

    query = query.replace('WHERE  AND', 'WHERE')

    print('Executing query: ' + query)

    cur.execute(query)

    data_row = cur.fetchone()

    print('Data for {region}, period: {period}, field: {field}, prediction: {prediction}'
          .format(region=reg.name, period=dt_range.description, field=query_field,
                  prediction=prediction.description()))

    print('Row count: {count}, AVG: {avg}, ABS AVG: {abs_avg}, STD DEVIATION: {std_dev}'
          .format(count=data_row[0], avg=data_row[1], abs_avg=data_row[2], std_dev=data_row[3]))

    return data_row


# Coordinates come from get_geo_params.sql
ALL_DATA = Coordinates(None, None, None, None, 'Wszystkie dane', 'Ten arkusz zawiera analizę'
                                                                 ' wszystkich dostępnych danych prognozowych z KSGmet. '
                                                                 'Analizowana jest różnica przewidzianej temperatury z '
                                                                 'temperaturą faktyczną (jako faktyczna brana jest '
                                                                 'prognoza na > 0h). Dane są brane z całego obszaru '
                                                                 'Polski. Pod-arkusze zawierają analizę danych dla '
                                                                 'wybranych obszarów geograficznych. Dane podzielone '
                                                                 'są pod względem na jak daleko wprzód w czasie jest '
                                                                 'prognoza, jak również na przedziały czasowe. '
                                                                 'Analizowane są 3 pola temperatury: t2mean2m, tmin2m '
                                                                 'i tmax2m. Pojedyncza komórka w pliku CSV traktowana '
                                                                 'jest jako osobna prognoza.', None)
POMORZE = Coordinates(131, 149, 145, 153, 'Pomorze', 'W tym pod-arkuszu analizowane są dane z prostokątnego obszaru, '
                                                     'gdzie przeciwległymi wierzchołkami są Bolszewo '
                                                     '(54.61688338418353, 18.17666407394085) i Cedry Małe '
                                                     '(54.27248258415296, 18.88028663455148).', 'pomorze.png')
MAZOWSZE = Coordinates(175, 228, 106, 117, 'Mazowsze', 'W tym pod-arkuszu analizowane są dane z prostokątnego obszaru, '
                                                       'gdzie przeciwległymi wierzchołkami są Lidzbark '
                                                       '(53.26058428867436, 19.823542973783265) i Ostów Mazowiecka '
                                                       '(52.79559391678067, 21.898069637271004).', 'mazowsze.png')
POLUDNIE = Coordinates(154, 178, 13, 38, 'Południe', 'W tym pod-arkuszu analizowane są dane z prostokątnego obszaru, '
                                                     'gdzie przeciwległymi wierzchołkami są Katowice (50.2515617199834,'
                                                     ' 19.035009483944236) i Zakopane (49.291731446084945, '
                                                     '19.985581547693716).', 'poludnie.png')

REGIONS = [ALL_DATA, POMORZE, MAZOWSZE, POLUDNIE]
DATE_RANGES = [DateRange(None, None, 'cały przedział czasowy'),
               DateRange(57, 122, '2018-10-27 do 2018-12-31'),
               DateRange(123, 181, '2019-01-01 do 2019-02-28'),
               DateRange(182, 213, '2019-03-01 do 2019-04-01')]
FIELDS = ['t2mean2m_delta', 'tmin2m_delta', 'tmax2m_delta']
PREDICTIONS = [Prediction(None, None),
               Prediction(4, 6),
               Prediction(7, 12),
               Prediction(13, 18),
               Prediction(19, 24),
               Prediction(25, 30),
               Prediction(31, 36),
               Prediction(37, 41)]

BOLD_FONT = Font(bold=True)

IMG_HEIGHT = 300
IMG_ROW_HEIGHT = 250
IMG_WIDTH = 400

connection = None
cursor = None

try:
    connection = psycopg2.connect(user='postgres',
                                  password='password',
                                  host='127.0.0.1',
                                  port='5432',
                                  database='ksgmet')
    cursor = connection.cursor()

    workbook = Workbook()
    sheet = None

    for region in REGIONS:
        if sheet:
            sheet = workbook.create_sheet()
        else:
            sheet = workbook.active

        sheet.title = region.name
        sheet.page_setup.fitToWidth = 1

        x, y = 1, 1

        sheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y + len(PREDICTIONS))
        desc_cell = sheet.cell(row=x, column=y, value=region.description)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

        row = sheet.row_dimensions[x]
        row.height = IMG_ROW_HEIGHT

        sheet.merge_cells(range_string='J1:L1')

        if region.map_file:
            img = Image('D:\\workspace\\MGR\\maps\\regions\\' + region.map_file)
            img.anchor = 'J1'
            img.width = IMG_WIDTH
            img.height = IMG_HEIGHT
            sheet.add_image(img)

        x += 1

        for field in FIELDS:

            sheet.cell(row=x, column=y, value='Pole ' + field)
            sheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y + 5)

            field_row = sheet.row_dimensions[x]
            field_row.font = BOLD_FONT
            field_row.fill = PatternFill('solid', fgColor='FFFF00')

            x += 1
            y = 1

            for date_range in DATE_RANGES:

                # Header
                header = '{} - {}'.format(field, date_range.description)
                header_cell = sheet.cell(row=x, column=y, value=header)
                sheet.merge_cells(start_row=x, start_column=y, end_row=x, end_column=y + len(PREDICTIONS))

                header_cell.font = BOLD_FONT
                header_cell.alignment = Alignment(horizontal='center')

                # Bold the column
                col = sheet.column_dimensions[get_column_letter(y)]
                col.width = 40

                x += 1

                cell = sheet.cell(row=x, column=y, value='Odległość prognozy')
                cell.font = BOLD_FONT
                x += 1

                cell = sheet.cell(row=x, column=y, value='Ilość analizowanych prognoz')
                cell.font = BOLD_FONT
                x += 1

                cell = sheet.cell(row=x, column=y, value='Średnia różnica')
                cell.font = BOLD_FONT
                x += 1

                cell = sheet.cell(row=x, column=y, value='Średnia różnica wartości bezwzględnych')
                cell.font = BOLD_FONT
                x += 1

                cell = sheet.cell(row=x, column=y, value='Odchylenie standardowe')
                cell.font = BOLD_FONT

                # Go to data coll
                x -= 4
                y += 1

                for prediction in PREDICTIONS:
                    cell = sheet.cell(row=x, column=y, value=prediction.description())
                    cell.font = BOLD_FONT
                    x += 1

                    data = execute_query(cursor, region, date_range, field, prediction)

                    for val in data:
                        sheet.cell(row=x, column=y, value=val)
                        x += 1

                    # Back up to data coll start and go to the next one
                    y += 1
                    x -= 5

                if region == ALL_DATA:
                    img_row = x + 5
                    img_col = y - len(PREDICTIONS) - 1

                    sheet.merge_cells(start_row=img_row, end_row=img_row,
                                      start_column=img_col, end_column=img_col + 2)

                    row = sheet.row_dimensions[img_row]
                    row.height = IMG_ROW_HEIGHT

                    img = Image('D:\\workspace\\MGR\\maps\\first_analysis\\' + date_range.map_dir_part() +
                                '\\' + field + '_avg_ad.png')
                    img.anchor = sheet.cell(row=img_row, column=img_col).coordinate
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    sheet.add_image(img)

                    img_col = img_col + 3

                    sheet.merge_cells(start_row=img_row, end_row=img_row,
                                      start_column=img_col, end_column=img_col + 5)

                    img = Image('D:\\workspace\\MGR\\maps\\first_analysis\\' + date_range.map_dir_part() +
                                '\\' + field + '_abs_avg_ad.png')
                    img.anchor = sheet.cell(row=img_row, column=img_col).coordinate
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    sheet.add_image(img)

                    img_col = img_col - 3
                    img_row = img_row + 1

                    row = sheet.row_dimensions[img_row]
                    row.height = IMG_ROW_HEIGHT

                    sheet.merge_cells(start_row=img_row, end_row=img_row,
                                      start_column=img_col, end_column=img_col + 2)

                    img = Image('D:\\workspace\\MGR\\maps\\first_analysis\\' + date_range.map_dir_part() +
                                '\\' + field + '_std_dev_ad.png')
                    img.anchor = sheet.cell(row=img_row, column=img_col).coordinate
                    img.width = IMG_WIDTH
                    img.height = IMG_HEIGHT
                    sheet.add_image(img)

                # Go back to header
                x -= 1
                # Space between date ranges
                col = sheet.column_dimensions[get_column_letter(y)]
                col.fill = PatternFill('solid', fgColor='E0E0E0')
                y += 1

            # New field, 8 = 5 data rows, 2 img row and one space
            x += 9
            y = 1

    workbook.save('D:\\workspace\\MGR\\data.xlsx')

finally:
    if cursor:
        cursor.close()
    if connection:
        connection.close()
