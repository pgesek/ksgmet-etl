from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font,  Alignment
from openpyxl.utils import get_column_letter


class XlsxDoc:

    BOLD_FONT = Font(bold=True)
    DESC_CELL_LENGTH = 10
    COL_WIDTH = 35

    def __init__(self):
        self.workbook = Workbook()
        self.sheet = None
        self.x = 1
        self.y = 1

    def create_sheet(self, title, description):
        if self.sheet:
            self.sheet = self.workbook.create_sheet(title=title)
        else:
            self.sheet = self.workbook.active
            self.sheet.title = title

        self.sheet.page_setup.fitToWidth = 1

        self.x = 1
        self.y = 1

        self.sheet.merge_cells(
            start_row=self.x,
            start_column=self.y,
            end_row=self.x,
            end_column=self.y + XlsxDoc.DESC_CELL_LENGTH
        )
        desc_cell = self._write_cell(description)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

        self._next_row()

    def write_table(self, data):
        title = data['title']
        col_headers = data['col_headers']
        row_headers = data['row_headers']
        content = data['content']

        self._next_row()

        self.sheet.merge_cells(
            start_row=self.x,
            start_column=self.y,
            end_row=self.x,
            end_column=self.y + len(col_headers)
        )
        title_cell = self._write_cell(title)
        title_cell.font = XlsxDoc.BOLD_FONT
        title_cell.fill = PatternFill('solid', fgColor='FFFF00')
        self._set_col_width()

        self._next_row()
        self._next_column()

        for col_header in col_headers:
            col_header_cell = self._write_cell(col_header)
            col_header_cell.font = XlsxDoc.BOLD_FONT

            self._set_col_width()

            self._next_column()

        self._first_column()

        for row_header in row_headers:
            row_header_cell = self._write_cell(row_header)
            row_header_cell.font = XlsxDoc.BOLD_FONT
            self._next_row()

        self._next_column()
        self._move_row(-(len(row_headers) - 1))

        for i in range(len(row_headers) - 1):
            for j in range(len(col_headers)):
                self._write_cell(content[i][j])
                self._next_column()

            self._next_row()
            self._first_column()
            self._next_column()

        self._first_column()

    def save(self, dest_path):
        self.workbook.save(dest_path)

    def _write_cell(self, value):
        return self.sheet.cell(row=self.x, column=self.y, value=value)

    def _next_row(self):
        self._move_row(1)

    def _move_row(self, step):
        self.x += step

    def _next_column(self):
        self._move_column(1)

    def _move_column(self, step):
        self.y += step

    def _first_column(self):
        self.y = 1

    def _set_col_width(self):
        col_dim = self.sheet.column_dimensions[get_column_letter(self.y)]
        col_dim.width = XlsxDoc.COL_WIDTH
