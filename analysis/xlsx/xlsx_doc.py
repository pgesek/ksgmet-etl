from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font,  Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


class XlsxDoc:

    BOLD_FONT = Font(bold=True)
    DESC_CELL_LENGTH = 2
    COL_WIDTH = 35

    IMG_NUM_OF_COLS = 2
    IMG_HEIGHT = 300
    IMG_ROW_HEIGHT = 250
    IMG_WIDTH = 400

    def __init__(self):
        self.workbook = Workbook()
        self.sheet = None
        self.x = 1
        self.y = 1

    def create_sheet(self, title, description):
        if self.sheet:
            self.sheet = self.workbook.create_sheet(title=title)
        else:
            self.sheet = self.workbook.active
            self.sheet.title = title

        self.sheet.page_setup.fitToWidth = 1

        self.x = 1
        self.y = 1

        self.sheet.merge_cells(
            start_row=self.x,
            start_column=self.y,
            end_row=self.x,
            end_column=self.y + XlsxDoc.DESC_CELL_LENGTH - 1
        )
        desc_cell = self._write_cell(description)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')

        self._move_column(self.DESC_CELL_LENGTH)

    def write_table(self, data):
        self._next_row()
        self._first_column()

        title = data['title']
        col_headers = data['col_headers']
        row_headers = data['row_headers']
        content = data['content']

        self._next_row()

        self.sheet.merge_cells(
            start_row=self.x,
            start_column=self.y,
            end_row=self.x,
            end_column=self.y + len(col_headers)
        )
        title_cell = self._write_cell(title)
        title_cell.font = XlsxDoc.BOLD_FONT
        title_cell.fill = PatternFill('solid', fgColor='FFFF00')
        self._set_col_width()

        self._next_row()
        self._next_column()

        for col_header in col_headers:
            col_header_cell = self._write_cell(col_header)
            col_header_cell.font = XlsxDoc.BOLD_FONT

            self._set_col_width()

            self._next_column()

        self._first_column()

        for row_header in row_headers:
            row_header_cell = self._write_cell(row_header)
            row_header_cell.font = XlsxDoc.BOLD_FONT
            self._next_row()

        self._next_column()
        self._move_row(-(len(row_headers) - 1))

        for i in range(len(row_headers) - 1):
            for j in range(len(col_headers)):
                content_cell = self._write_cell(content[i][j])
                content_cell.alignment = Alignment(horizontal='left')

                self._next_column()

            self._next_row()
            self._first_column()
            self._next_column()

        self._first_column()
        self._next_row()

    def write_image(
            self,
            img_path,
            img_width=IMG_WIDTH,
            img_height=IMG_HEIGHT,
            img_row_height=IMG_ROW_HEIGHT,
            img_num_of_cols=IMG_NUM_OF_COLS
    ):
        self.sheet.merge_cells(
            start_row=self.x,
            end_row=self.x,
            start_column=self.y,
            end_column=self.y + img_num_of_cols - 1
        )

        img_row = self.sheet.row_dimensions[self.x]
        img_row.height = img_row_height

        img = Image(img_path)

        img.anchor = self.sheet.cell(
            row=self.x,
            column=self.y
        ).coordinate

        img.width = img_width
        img.height = img_height

        self.sheet.add_image(img)
        self._move_column(XlsxDoc.IMG_NUM_OF_COLS)

    def save(self, dest_path):
        self.workbook.save(dest_path)

    def _write_cell(self, value):
        return self.sheet.cell(row=self.x, column=self.y, value=value)

    def _next_row(self):
        self._move_row(1)

    def _move_row(self, step):
        self.x += step

    def _next_column(self):
        self._move_column(1)

    def _move_column(self, step):
        self.y += step

    def _first_column(self):
        self.y = 1

    def _set_col_width(self):
        col_dim = self.sheet.column_dimensions[get_column_letter(self.y)]
        col_dim.width = XlsxDoc.COL_WIDTH
